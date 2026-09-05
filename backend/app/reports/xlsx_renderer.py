"""XLSX Report Renderer using openpyxl."""

import io
from datetime import datetime, timezone
from typing import Any, Dict, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def sanitize_xlsx_value(val: Any) -> Any:
    if val is None:
        return ""
    if isinstance(val, str):
        if val.startswith(("=", "+", "-", "@")):
            return "'" + val
        return val
    return val


class XLSXReportRenderer:
    @staticmethod
    def render_report(report_type: str, data: Dict[str, Any], title: str) -> bytes:
        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Styling definitions
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
        title_font = Font(name="Calibri", size=14, bold=True, color="1E293B")
        sub_font = Font(name="Calibri", size=9, italic=True, color="64748B")
        thin_border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )

        # Title Block
        ws_summary.cell(row=1, column=1, value=f"DealFlow360 — {title}").font = title_font
        gen_time = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        ws_summary.cell(row=2, column=1, value=f"Generated At: {gen_time} | Confidential").font = sub_font

        # Summary KPIs
        ws_summary.cell(row=4, column=1, value="Metric").font = header_font
        ws_summary.cell(row=4, column=1).fill = header_fill
        ws_summary.cell(row=4, column=2, value="Value").font = header_font
        ws_summary.cell(row=4, column=2).fill = header_fill

        current_row = 5
        for k, v in data.items():
            if isinstance(v, (int, float, str)) and k not in ["start_date", "end_date", "granularity"]:
                ws_summary.cell(row=current_row, column=1, value=sanitize_xlsx_value(k.replace("_", " ").title())).border = thin_border
                ws_summary.cell(row=current_row, column=2, value=sanitize_xlsx_value(v)).border = thin_border
                current_row += 1
            elif isinstance(v, dict) and all(isinstance(val, (int, float, str)) for val in v.values()):
                val_str = ", ".join(f"{curr}: {val}" for curr, val in v.items())
                ws_summary.cell(row=current_row, column=1, value=sanitize_xlsx_value(k.replace("_", " ").title())).border = thin_border
                ws_summary.cell(row=current_row, column=2, value=sanitize_xlsx_value(val_str)).border = thin_border
                current_row += 1
            elif isinstance(v, list) and v and isinstance(v[0], dict):
                # Add extra sheet for list datasets
                sheet_title = k.replace("_", " ").title()[:30]
                ws_list = wb.create_sheet(title=sheet_title)

                keys = list(v[0].keys())
                for col_idx, col_name in enumerate(keys, start=1):
                    c = ws_list.cell(row=1, column=col_idx, value=sanitize_xlsx_value(col_name.replace("_", " ").title()))
                    c.font = header_font
                    c.fill = header_fill

                for r_idx, row_dict in enumerate(v, start=2):
                    for c_idx, col_name in enumerate(keys, start=1):
                        cell_val = row_dict.get(col_name)
                        if isinstance(cell_val, dict):
                            cell_val = ", ".join(f"{ck}: {cv}" for ck, cv in cell_val.items())
                        ws_list.cell(row=r_idx, column=c_idx, value=sanitize_xlsx_value(cell_val)).border = thin_border

                for col in ws_list.columns:
                    max_len = max(len(str(cell.value or "")) for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    ws_list.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # Autofit summary columns
        for col in ws_summary.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_summary.column_dimensions[col_letter].width = max(max_len + 4, 15)

        buffer = io.BytesIO()
        wb.save(buffer)
        xlsx_bytes = buffer.getvalue()
        buffer.close()
        return xlsx_bytes
